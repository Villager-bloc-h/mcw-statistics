import json
import os
import time
from datetime import datetime

import openpyxl

import base

params = {
    "action": "query",
    "format": "json",
    "list": "logevents",
    "formatversion": 2,
    "leprop": "user",
    "letype": "patrol/patrol",
    "lelimit": "max",
}

user_list = {}
last_lecontinue = ""
loop_count = 0
current_time = ""

wb = openpyxl.Workbook()
ws = wb.active
ws['A1'] = "排名"
ws['B1'] = "用户名"
ws['C1'] = "巡查数"
ws.column_dimensions['B'].width = 20.00

if os.path.exists("patrolcount_backup.json"): # 恢复上次中断时保存的内容
    with open(f"patrolcount_backup.json", "r", encoding="utf-8") as backup_file:
        backup_data = json.load(backup_file)

    last_lecontinue = backup_data["last_lecontinue"]
    loop_count = backup_data["loop_count"]
    user_list = backup_data["user_list"]
    current_time = backup_data["current_time"]

    print("已恢复上次中断时保存的内容")

if current_time == "":
    current_time = datetime.now().strftime("%Y%m%d%H%M%S")

excel_filename = f"patrolcount-{current_time}.xlsx"
txt_filename = f"patrolcount-{current_time}.txt"

print("启动成功", end='\n\n')

while True: # 获取所有巡查日志的内容
    time.sleep(3)

    if last_lecontinue != "":  # 不是首次循环，使用这个继续
        last_params = params.copy()
        last_params.update({"lecontinue": last_lecontinue})
    else:  # 首次循环
        last_params = params

    le_data = base.get_data(last_params)

    loop_count += 1
    print(f"成功获取第{loop_count}组数据")

    for item in le_data['query']['logevents']:
        username = item["user"]
        # 更新计数：存在则+1，不存在则初始化为1
        user_list[username] = user_list.get(username, 0) + 1

    if 'continue' not in le_data:  # 已经全部获取完成，跳出循环
        break

    last_lecontinue = le_data['continue']['lecontinue']

    if loop_count % 10 == 0:
        backup_json = {
            "last_lecontinue": last_lecontinue,
            "loop_count": loop_count,
            "current_time": current_time,
            "user_list": user_list,
        }
        with open(f'patrolcount_backup.json', 'w', encoding='utf-8') as file:
            json.dump(backup_json, file, ensure_ascii=False, indent=4)
        print("已保存进度")

print("所有数据已经获取完毕，正在处理中...")

sorted_data = []

# 将巡查次数信息放入列表
for user, count in user_list.items():
    sorted_data.append((user, count))

sorted_data.sort(key=lambda x: x[1], reverse=True)

ranks = []
current_rank = 1
prev_count = sorted_data[0][1]

# 添加排名数据
for i, data in enumerate(sorted_data):
    count = data[1]
    if count != prev_count:
        current_rank = i + 1
    ranks.append(current_rank)
    prev_count = count

# 将排序后的数据写入工作表
for idx, (user, count) in enumerate(sorted_data):
    row_idx = idx + 2
    ws.cell(row=row_idx, column=1, value=ranks[idx] if ranks else idx + 1)
    ws.cell(row=row_idx, column=2, value=user)
    ws.cell(row=row_idx, column=3, value=count)

wb.save(excel_filename)
print(f"Excel结果已保存至{excel_filename}")

if os.path.exists("patrolcount_backup.json"):
    os.remove("patrolcount_backup.json")
    print("已删除JSON备份文件")

# 将排序后的内容变为wikitable
wiki_content = '''{| class="wikitable sortable collapsible"'''


wiki_content = wiki_content + '''
|-
! 排名 !! 用户名 !! 巡查数
'''

for idx, (user, count) in enumerate(sorted_data):
    
    rank = ranks[idx] if ranks else idx + 1
    wiki_content += "|-\n"
    wiki_content += f"| {rank} || [[User:{user}|{user}]] || {count}\n"
   
wiki_content += "|}"

# 将wikitable写入文本文件
with open(txt_filename, "w", encoding="utf-8") as f:
    f.write(wiki_content)
print(f"Wiki表格已保存至{txt_filename}")

input("按任意键退出")


