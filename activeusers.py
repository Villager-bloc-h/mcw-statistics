import time
from datetime import datetime, timedelta, timezone
import openpyxl

import base

# 读取配置，包括：用户组排列顺序、用户组名称和模式
usergroup_order, usergroup_mapping, mode = base.activeusers_get_config()

now = datetime.now()
year = now.year
month = now.month

if mode == "standard": # 标准模式
    day1 = 28
    day2 = 28
    if month == 2:
        day2 = 25
    elif month == 3:
        day1 = 25

    end_time = now.replace(day=day2, hour=0, minute=0, second=0)
    start_time = end_time.replace(day=day1, month=month - 1)
    if start_time.month == 0:
        start_time = start_time.replace(year=year - 1, month=12)

    start_date_str = f"{start_time.year}年{start_time.month}月{start_time.day}日0时"
    end_date_str = f"{end_time.year}年{end_time.month}月{end_time.day}日0时"

    excel_filename = f"activeusers-{year}年{month}月.xlsx"
    txt_filename = f"activeusers-{year}年{month}月.txt"

elif mode == "debug": # 调试模式
    end_time = now.replace(second=0)
    start_time = end_time - timedelta(days=30)

    start_date_str = f"{start_time.year}年{start_time.month}月{start_time.day}日{start_time.hour}时{start_time.minute}分"
    end_date_str = f"{end_time.year}年{end_time.month}月{end_time.day}日{end_time.hour}时{end_time.minute}分"

    excel_filename = end_time.strftime("activeusers-%Y年%m月%d日%H时%M分-调试.xlsx")
    txt_filename = end_time.strftime("activeusers-%Y年%m月%d日%H时%M分-调试.txt")

# API只接受UTC时间的时间戳，因此需要根据时区做合适处理
end_time = end_time.replace(tzinfo=timezone(timedelta(hours=base.timezone)))
start_time = start_time.replace(tzinfo=timezone(timedelta(hours=base.timezone)))
end_time_utc = end_time.astimezone(timezone.utc)
start_time_utc = start_time.astimezone(timezone.utc)
end_timestamp_utc = end_time_utc.strftime("%Y-%m-%dT%H:%M:%SZ")
start_timestamp_utc = start_time_utc.strftime("%Y-%m-%dT%H:%M:%SZ")

params = {
    "action": "query",
    "format": "json",
    "list": "recentchanges",
    "formatversion": 2,
    "rcprop": "user|loginfo",
    "rclimit": "max",
    "rctype": "edit|new|log",
}
params.update({
    "rcstart": end_timestamp_utc,
    "rcend": start_timestamp_utc,
})

# 设置excel表格格式
wb = openpyxl.Workbook()
ws = wb.active
ws['A1'] = "排名"
ws['B1'] = "用户名"
ws['C1'] = "操作数"
ws['D1'] = "本地用户组"
ws.column_dimensions['B'].width = 20.00
ws.column_dimensions['D'].width = 10.89

user_list = {}
last_rccontinue = ""
loop_count = 0

print("启动成功", end='\n\n')

while True: # 获取过去一个月的最近更改详情
    time.sleep(3)

    if last_rccontinue != "": # 不是首次循环，使用这个继续
        last_params = params.copy()
        last_params.update({"rccontinue": last_rccontinue})
    else: # 首次循环
        last_params = params

    rc_data = base.get_data(last_params)

    loop_count += 1
    print(f"成功获取第{loop_count}组数据")

    for item in rc_data['query']['recentchanges']:
        if item['type'] == "log":
            if 'actionhidden' in item:
                if 'userhidden' in item: # 用户名已移除，忽略
                    continue
            elif item['logtype'] == "newusers": # 用户创建日志不计入操作数
                continue

        username = item["user"]
        # 更新计数：存在则+1，不存在则初始化为1
        user_list[username] = user_list.get(username, 0) + 1

    if 'continue' not in rc_data: # 已经全部获取完成，跳出循环
        break

    last_rccontinue = rc_data['continue']['rccontinue']

print("所有数据已经获取完毕，正在处理中...")

# 获取用户组信息
usergroups_params = {
    "action": "query",
    "format": "json",
    "list": "allusers",
    "formatversion": 2,
    "augroup": "autopatrol|bot|bureaucrat|interface-admin|patrollers|sysop",
    "auprop": "groups",
    "aulimit": "max",
}
usergroups_data = base.get_data(usergroups_params)

user_permissions = {}

# 置为最高级别用户组
for user in usergroups_data['query']['allusers']:
    username = user['name']
    groups = set(user['groups'])

    highest_perm = None
    for perm in usergroup_order:
        if perm in groups:
            highest_perm = perm
            break

    user_permissions[username] = highest_perm

sorted_data = []

# 过滤IP用户，将用户组信息放入列表
for user, count in user_list.items():
    if base.is_ip_address(user):
        continue

    # 获取用户组信息
    usergroup = user_permissions.get(user)
    group_name = usergroup_mapping[usergroup] if usergroup else "无"

    sorted_data.append((user, count, group_name))

sorted_data.sort(key=lambda x: x[1], reverse=True)

ranks = []
current_rank = 1
prev_count = sorted_data[0][1]

# 添加排名数据
for i, data in enumerate(sorted_data):
    count = data[1]
    if count != prev_count:
        current_rank = i + 1
    ranks.append(current_rank)
    prev_count = count

# 将排序后的数据写入工作表
for idx, (user, count, group_name) in enumerate(sorted_data):
    row_idx = idx + 2
    ws.cell(row=row_idx, column=1, value=ranks[idx] if ranks else idx + 1)
    ws.cell(row=row_idx, column=2, value=user)
    ws.cell(row=row_idx, column=3, value=count)
    ws.cell(row=row_idx, column=4, value=group_name)

wb.save(excel_filename)
print(f"Excel结果已保存至{excel_filename}")

# 将排序后的内容变为wikitable
wiki_content = '''{| class="wikitable sortable collapsible"
|+ %s-%s''' % (start_date_str, end_date_str)

if base.timezone > 0:
    wiki_content = f"{wiki_content} (UTC+{base.timezone}) "
elif base.timezone < 0:
    wiki_content = f"{wiki_content} (UTC-{base.timezone}) "
else:
    wiki_content = f"{wiki_content} (UTC) "

wiki_content = wiki_content + '''活跃用户列表
|-
! 排名 !! 用户名 !! 操作数 !! 本地用户组
'''

for idx, (user, count, group_name) in enumerate(sorted_data):
    group_display = group_name
    if group_name != "无":
        group_display = f"[[Minecraft Wiki:{group_name}|{group_name}]]"

    rank = ranks[idx] if ranks else idx + 1
    wiki_content += "|-\n"
    wiki_content += f"| {rank} || [[User:{user}|{user}]] || {count} || {group_display}\n"

wiki_content += "|}"

# 将wikitable写入文本文件
with open(txt_filename, "w", encoding="utf-8") as f:
    f.write(wiki_content)
print(f"Wiki表格已保存至{txt_filename}")

'''
# 可能是未来要加入的内容
params.update({
        "rcprop": "user|loginfo|title|sizes|flags",
    })
活跃用户分数量化指标
* 操作数x100
* 增减字节数绝对值之和x0.01
* 创建新页面次数x1000
* 在内容命名空间进行编辑x500
* 被封禁1次-10000
'''
