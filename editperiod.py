import openpyxl
import datetime

import base

datafile, contribs_data = base.editperiod_get_config()

username = contribs_data[0]["user"]

wb = openpyxl.Workbook()

# 第一个工作表
sheet1 = wb.active
sheet1.title = "按年计"
sheet1['A1'] = "年份"
sheet1['B1'] = "编辑数"

# 第二个工作表
sheet2 = wb.create_sheet("按1-12月计")
sheet2['A1'] = "月份"
sheet2['B1'] = "编辑数"

row = 2
months = ["1月","2月","3月","4月","5月","6月","7月","8月","9月","10月","11月","12月"]

while row < 14: # 第二个表格内容初始化
    sheet2[f'A{row}'] = months[row-2]
    sheet2[f'B{row}'] = 0
    row += 1

# 第三个工作表
sheet3 = wb.create_sheet("按一周七天计")
sheet3['A1'] = "星期"
sheet3['B1'] = "编辑数"

row = 2
weekdays = ["星期一","星期二","星期三","星期四","星期五","星期六","星期日"]

while row < 9: # 第三个表格内容初始化
    sheet3[f'A{row}'] = weekdays[row-2]
    sheet3[f'B{row}'] = 0
    row += 1

# 第四个工作表
sheet4 = wb.create_sheet("按1-31日计")
sheet4['A1'] = "天数"
sheet4['B1'] = "编辑数"

row = 2
days = list(range(1, 32))

while row < 33: # 第四个表格内容初始化
    sheet4[f'A{row}'] = days[row-2]
    sheet4[f'B{row}'] = 0
    row += 1

# 第五个工作表
sheet5 = wb.create_sheet("按每小时计")
sheet5['A1'] = "时间段"
sheet5['B1'] = "编辑数"
sheet5.column_dimensions['A'].width = 12.00

row = 2
hours = ["00:00-01:00","01:00-02:00","02:00-03:00","03:00-04:00","04:00-05:00","05:00-06:00","06:00-07:00","07:00-08:00","08:00-09:00","09:00-10:00","10:00-11:00","11:00-12:00","12:00-13:00","13:00-14:00","14:00-15:00","15:00-16:00","16:00-17:00","17:00-18:00","18:00-19:00","19:00-20:00","20:00-21:00","21:00-22:00","22:00-23:00","23:00-00:00"]

while row < 26: # 第五个表格内容初始化
    sheet5[f'A{row}'] = hours[row-2]
    sheet5[f'B{row}'] = 0
    row += 1

# 第六个工作表
sheet6 = wb.create_sheet("按每分钟计")
sheet6['A1'] = "分钟数"
sheet6['B1'] = "编辑数"
sheet6.column_dimensions['A'].width = 12.00

row = 2
minutes = list(range(0, 60))

while row < 62: # 第六个表格内容初始化
    sheet6[f'A{row}'] = minutes[row-2]
    sheet6[f'B{row}'] = 0
    row += 1

# 第七个工作表
sheet7 = wb.create_sheet("按每秒钟计")
sheet7['A1'] = "秒钟数"
sheet7['B1'] = "编辑数"
sheet7.column_dimensions['A'].width = 12.00

row = 2
seconds = minutes

while row < 62: # 第七个表格内容初始化
    sheet7[f'A{row}'] = seconds[row-2]
    sheet7[f'B{row}'] = 0
    row += 1

# 第八个工作表
sheet8 = wb.create_sheet("按年月计")
sheet8['A1'] = "年月"
sheet8['B1'] = "编辑数"
sheet8.column_dimensions['A'].width = 12.00

# 第九个工作表
sheet9 = wb.create_sheet("按年月日计")
sheet9['A1'] = "年月日"
sheet9['B1'] = "编辑数"
sheet9.column_dimensions['A'].width = 12.00

year_dict = {}
year_month_dict = {}
year_month_day_dict = {}
min_date = None
max_date = None

for item in contribs_data:
    dt = base.extract_from_timestamp(item["timestamp"])

    year = dt.year
    year_dict[year] = year_dict.get(year, 0) + 1

    year_month_key = f"{dt.year}-{dt.month:02d}"
    year_month_dict[year_month_key] = year_month_dict.get(year_month_key, 0) + 1

    year_month_day_key = f"{dt.year}-{dt.month:02d}-{dt.day:02d}"
    year_month_day_dict[year_month_day_key] = year_month_day_dict.get(year_month_day_key, 0) + 1

    d_only = dt.date()
    if min_date is None or d_only < min_date:
        min_date = d_only
    if max_date is None or d_only > max_date:
        max_date = d_only

    month = dt.month
    sheet2.cell(row=month+1, column=2).value += 1

    weekday = dt.weekday()
    sheet3.cell(row=weekday+2, column=2).value += 1

    day = dt.day
    sheet4.cell(row=day+1, column=2).value += 1

    hour = dt.hour
    sheet5.cell(row=hour+2, column=2).value += 1

    minute = dt.minute
    sheet6.cell(row=minute+2, column=2).value += 1

    second = dt.second
    sheet7.cell(row=second+2, column=2).value += 1

row = 2
for y in sorted(year_dict):
    sheet1[f'A{row}'] = y
    sheet1[f'B{row}'] = year_dict[y]
    row += 1

row = 2
for ym in sorted(year_month_dict):
    sheet8[f'A{row}'] = ym
    sheet8[f'B{row}'] = year_month_dict[ym]
    row += 1

row = 2
if min_date is not None and max_date is not None:
    cur = min_date
    one_day = datetime.timedelta(days=1)
    while cur <= max_date:
        ymd = cur.strftime("%Y-%m-%d")
        sheet9[f'A{row}'] = ymd
        sheet9[f'B{row}'] = year_month_day_dict.get(ymd, 0)
        row += 1
        cur += one_day

current_time = datafile[-14:]
base.output(f"{username}-editperiod-{current_time}.xlsx", wb, "xlsx")

print(f"结果已保存至{username}-editperiod-{current_time}.xlsx")
