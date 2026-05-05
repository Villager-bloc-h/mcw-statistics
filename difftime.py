import os
import time
import openpyxl
from datetime import datetime

from openpyxl.styles import NamedStyle

import base

params = {
    "action": "compare",
    "format": "json",
    "formatversion": 2,
    "prop": "timestamp",
}
num_diff, minrev, maxrev = base.difftime_get_config()

if minrev is None:
    minrev = 1
elif minrev < 1:
    print("minrev超出范围，已重置为1")
    minrev = 1

base.login()

current_maxrev = base.get_last_diff()['query']['recentchanges'][0]['revid']

if maxrev is None:
    maxrev = current_maxrev
elif maxrev > current_maxrev:
    print("maxrev超出范围，已重置为当前最大revid")
    maxrev = current_maxrev

if minrev > maxrev:
    print("minrev大于maxrev，已重置为1和当前最大revid")
    minrev = 1
    maxrev = current_maxrev

rev = minrev

if os.path.exists("difftime_backup.xlsx"): # 恢复上次中断时保存的内容
    wb = openpyxl.load_workbook("difftime_backup.xlsx")
    ws = wb.active
    date_style = NamedStyle(name="date_style", number_format="YYYY-MM-DD")
    wb.add_named_style(date_style)
    last_rev = ws.cell(row=ws.max_row, column=2).value
    rev = last_rev + num_diff
else:
    wb = openpyxl.Workbook()
    ws = wb.active
    date_style = NamedStyle(name="date_style", number_format="YYYY-MM-DD")
    wb.add_named_style(date_style)
    ws.column_dimensions['A'].width = 10.89
    ws.column_dimensions['C'].width = 10.89
    ws['A1'] = "日期"
    ws['B1'] = "revid"

fromrev = rev
torev = rev

current_time = datetime.now().strftime("%Y%m%d%H%M%S")

print("启动成功", end='\n\n')

while rev < maxrev: # 主循环
    time.sleep(1)

    current_params = params.copy()
    current_params.update({
        "fromrev": fromrev,
        "torev": torev,
    })

    data = base.get_data(current_params)

    if 'compare' in data:
        row = ws.max_row + 1
        fromtimestamp = data['compare']['fromtimestamp']
        totimestamp = data['compare']['totimestamp']
        dt1 = base.extract_from_timestamp(fromtimestamp)
        dt2 = base.extract_from_timestamp(totimestamp)
        year1, month1, day1 = dt1.year, dt1.month, dt1.day
        year2, month2, day2 = dt2.year, dt2.month, dt2.day
        from_date_obj = datetime(year1, month1, day1)

        if year1 != year2 or month1 != month2 or day1 != day2:
            to_date_obj = datetime(year2, month2, day2)
            ws.cell(row=row, column=3, value=to_date_obj).style = date_style

        ws.cell(row=row, column=1, value=from_date_obj).style = date_style
        ws.cell(row=row, column=2, value=fromrev)

        print(f"成功获取revid为{fromrev}的差异数据")

        rev = rev + num_diff
        fromrev = rev
        torev = rev

        if row % 100 == 0:
            wb.save("difftime_backup.xlsx")
            print(f"已保存进度至revid：{fromrev}")

    elif rev - fromrev > 5 or torev - rev > 5: # 防止大量修订数据被删除时拖慢程序速度
        print(f"{fromrev}-{torev}的差异数据均无法查询！")

        rev = rev + num_diff
        fromrev = rev
        torev = rev

    else: # 修订数据被删除，扩大搜索范围
        fromrev_params = params.copy()
        current_params.update({
            "fromrev": fromrev,
            "torev": fromrev,
        })
        torev_params = params.copy()
        torev_params.update({
            "fromrev": torev,
            "torev": torev,
        })

        fromrev_data = base.get_data(fromrev_params)
        time.sleep(0.5)
        torev_data = base.get_data(torev_params)

        if 'error' in fromrev_data and fromrev > minrev:
            fromrev = fromrev - 1

        if 'error' in torev_data and torev < maxrev:
            torev = torev + 1

base.output(f"difftime-{current_time}.xlsx", wb, "xlsx")
print(f"结果已保存至difftime-{current_time}.xlsx")

if os.path.exists("difftime_backup.xlsx"):
    os.remove("difftime_backup.xlsx")
    print("已删除Excel备份文件")
