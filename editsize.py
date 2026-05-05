import openpyxl

import base

datafile, contribs_data = base.editperiod_get_config()

username = contribs_data[0]["user"]

wb = openpyxl.Workbook()

# 第一个工作表
sheet1 = wb.active
sheet1.title = "1000"
sheet1['A1'] = "字节更改范围"
sheet1['B1'] = "编辑数"
sheet1.column_dimensions['A'].width = 20.00

# 第二个工作表
sheet2 = wb.create_sheet("100")
sheet2['A1'] = "字节更改范围"
sheet2['B1'] = "编辑数"
sheet2.column_dimensions['A'].width = 20.00

# 第三个工作表
sheet3 = wb.create_sheet("10")
sheet3['A1'] = "字节更改范围"
sheet3['B1'] = "编辑数"
sheet3.column_dimensions['A'].width = 20.00

# 第四个工作表
sheet4 = wb.create_sheet("1")
sheet4['A1'] = "字节更改范围"
sheet4['B1'] = "编辑数"
sheet4.column_dimensions['A'].width = 20.00

# 初始化计数器
bin_1   = {}
bin_10  = {}
bin_100 = {}
bin_1000 = {}

for item in contribs_data:
    diff = item["sizediff"]

    # 区间宽度 1
    key_1 = str(diff)
    bin_1[key_1] = bin_1.get(key_1, 0) + 1

    # 区间宽度 10
    lower_10 = ((diff + 5) // 10) * 10 - 5
    key_10 = f"{lower_10},{lower_10 + 9}"
    bin_10[key_10] = bin_10.get(key_10, 0) + 1

    # 区间宽度 100
    lower_100 = ((diff + 50) // 100) * 100 - 50
    key_100 = f"{lower_100},{lower_100 + 99}"
    bin_100[key_100] = bin_100.get(key_100, 0) + 1

    # 区间宽度 1000
    lower_1000 = ((diff + 500) // 1000) * 1000 - 500
    key_1000 = f"{lower_1000},{lower_1000 + 999}"
    bin_1000[key_1000] = bin_1000.get(key_1000, 0) + 1

# 批量写入工作表
def write_sheet(sheet, counter):
    row = 2
    for k in sorted(counter.keys(), key=lambda x: int(x.split(",")[0])):
        sheet[f'A{row}'] = k
        sheet[f'B{row}'] = counter[k]
        row += 1

write_sheet(sheet4, bin_1)
write_sheet(sheet3, bin_10)
write_sheet(sheet2, bin_100)
write_sheet(sheet1, bin_1000)

current_time = datafile[-14:]
base.output(f"{username}-editsize-{current_time}.xlsx", wb, "xlsx")

print(f"结果已保存至{username}-editsize-{current_time}.xlsx")
