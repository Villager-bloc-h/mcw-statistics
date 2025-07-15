import json
import os
import time
from datetime import datetime

import openpyxl

import base

api_url = base.WIKI_API_URL + "?action=query&format=json&list=allrevisions&formatversion=2&arvprop=user&arvlimit=500"

user_list = {}
last_arvcontinue = ""
loop_count = 0
hidden = 0
current_time = ""

wb = openpyxl.Workbook()
ws = wb.active
ws['A1'] = "排名"
ws['B1'] = "用户名"
ws['C1'] = "编辑数"
ws.column_dimensions['B'].width = 20.00

if os.path.exists("editcount_backup.json"): # 恢复上次中断时保存的内容
    with open(f"editcount_backup.json", "r", encoding="utf-8") as backup_file:
        backup_data = json.load(backup_file)

    last_arvcontinue = backup_data["last_arvcontinue"]
    loop_count = backup_data["loop_count"]
    hidden = backup_data["hidden"]
    user_list = backup_data["user_list"]
    current_time = backup_data["current_time"]

    print("已恢复上次中断时保存的内容")

if current_time == "":
    current_time = datetime.now().strftime("%Y%m%d%H%M%S")

excel_filename = f"editcount-{current_time}.xlsx"

print("启动成功", end='\n\n')

while True: # 获取所有修订版本的内容
    time.sleep(3)

    if last_arvcontinue != "":  # 不是首次循环，使用这个继续
        last_api_url = api_url + "&arvcontinue=" + last_arvcontinue
    else:  # 首次循环
        last_api_url = api_url

    arv_data = base.get_data(last_api_url)

    loop_count += 1
    print(f"成功获取第{loop_count}组数据")

    for page in arv_data['query']['allrevisions']:
        for rev in page['revisions']:
            username = rev.get('user')
            if username is None:
                hidden += 1
                continue
            user_list[username] = user_list.get(username, 0) + 1

    if 'continue' not in arv_data:  # 已经全部获取完成，跳出循环
        break

    last_arvcontinue = arv_data['continue']['arvcontinue']

    if loop_count % 10 == 0:
        backup_json = {
            "last_arvcontinue": last_arvcontinue,
            "loop_count": loop_count,
            "hidden": hidden,
            "current_time": current_time,
            "user_list": user_list,
        }
        with open(f'editcount_backup.json', 'w', encoding='utf-8') as file:
            json.dump(backup_json, file, ensure_ascii=False, indent=4)
        print("已保存进度")

print("所有数据已经获取完毕，正在处理中...")

sorted_data = []

# 将编辑次数信息放入列表
for user, count in user_list.items():
    sorted_data.append((user, count))

sorted_data.sort(key=lambda x: x[1], reverse=True)

ranks = []
current_rank = 1
prev_count = sorted_data[0][1]

# 添加排名数据
for i, data in enumerate(sorted_data):
    count = data[1]
    if count != prev_count:
        current_rank = i + 1
    ranks.append(current_rank)
    prev_count = count

# 将排序后的数据写入工作表
for idx, (user, count) in enumerate(sorted_data):
    row_idx = idx + 2
    ws.cell(row=row_idx, column=1, value=ranks[idx] if ranks else idx + 1)
    ws.cell(row=row_idx, column=2, value=user)
    ws.cell(row=row_idx, column=3, value=count)

if hidden:
    current_row = ws.max_row + 1
    ws.cell(row=current_row, column=2, value="（用户名已移除）")
    ws.cell(row=current_row, column=3, value=hidden)

wb.save(excel_filename)
print(f"Excel结果已保存至{excel_filename}")

TOP_N = 1000
if len(sorted_data) > TOP_N:
    top_filename = f"editcount-{TOP_N}-{current_time}.xlsx"
    wb_top = openpyxl.Workbook()
    ws_top = wb_top.active
    ws_top['A1'] = "排名"
    ws_top['B1'] = "用户名"
    ws_top['C1'] = "编辑数"
    ws_top.column_dimensions['B'].width = 20.00

    for idx, (user, count) in enumerate(sorted_data[:TOP_N]):
        row_idx = idx + 2
        ws_top.cell(row=row_idx, column=1, value=ranks[idx] if ranks else idx + 1)
        ws_top.cell(row=row_idx, column=2, value=user)
        ws_top.cell(row=row_idx, column=3, value=count)

    if hidden:
        current_row = ws_top.max_row + 1
        ws_top.cell(row=current_row, column=2, value="（用户名已移除）")
        ws_top.cell(row=current_row, column=3, value=hidden)

    wb_top.save(top_filename)
    print(f"前{TOP_N}名已保存至{top_filename}")

if os.path.exists("editcount_backup.json"):
    os.remove("editcount_backup.json")
    print("已删除JSON备份文件")

input("按任意键退出")
