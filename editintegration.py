from datetime import timedelta

import openpyxl
from openpyxl.styles import NamedStyle
from openpyxl.chart import LineChart, Reference

import base


# 从editnamespace.py中提取的命名空间显示名称逻辑
def get_ns_display(title, ns):
    if ns == 0:
        return '（主）'
    if ':' in title:
        ns_name = title.split(':', 1)[0]
        # 特殊处理：Minecraft前缀视为主命名空间
        return '（主）' if ns_name == 'Minecraft' else ns_name
    return f"命名空间{ns}"


datafile, contribs_data = base.editperiod_get_config()

username = contribs_data[0]["user"]

# 收集所有命名空间显示名称
ns_display_names = set()
for rec in contribs_data:
    title = rec["title"]
    ns = rec["ns"]
    ns_display = get_ns_display(title, ns)
    ns_display_names.add(ns_display)

# 排序命名空间：主命名空间排最前，其余按字母顺序
sorted_ns = sorted(ns_display_names, key=lambda x: (x != '（主）', x))

# 初始化数据结构（添加百分比样式）
wb = openpyxl.Workbook()
ws = wb.active
date_style = NamedStyle(name="date_style", number_format="YYYY-MM-DD")
percent_style = NamedStyle(name="percent_style", number_format="0.00%")  # 新增百分比样式
wb.add_named_style(date_style)
wb.add_named_style(percent_style)  # 注册百分比样式
ws.column_dimensions['A'].width = 10.89

# 修改列标题（添加百分比标识）
columns = ["日期"] + [f"{ns}" for ns in sorted_ns] + ["总编辑数"]  # 修改列标题

# 写入列标题
for col_idx, col_name in enumerate(columns, 1):
    ws.cell(row=1, column=col_idx, value=col_name)

start_timestamp = base.extract_from_timestamp(contribs_data[0]["timestamp"])
end_timestamp = base.extract_from_timestamp(contribs_data[-1]["timestamp"])

# 按日期和命名空间统计编辑数
daily_edits = {}
for item in contribs_data:
    date_str = base.extract_from_timestamp(item["timestamp"]).strftime("%Y-%m-%d")
    title = item["title"]
    ns = item["ns"]
    ns_display = get_ns_display(title, ns)

    if date_str not in daily_edits:
        daily_edits[date_str] = {}

    daily_edits[date_str][ns_display] = daily_edits[date_str].get(ns_display, 0) + 1
    daily_edits[date_str]["总编辑数"] = daily_edits[date_str].get("总编辑数", 0) + 1

# 初始化累计计数器
accumulative_counts = {ns: 0 for ns in sorted_ns}
accumulative_total = 0  # 总编辑数的累计值

# 填充表格数据
row = 2
cur_date = start_timestamp
one_day = timedelta(days=1)

while cur_date <= end_timestamp:
    date_str = cur_date.strftime("%Y-%m-%d")
    ws.cell(row=row, column=1, value=cur_date).style = "date_style"

    day_data = daily_edits.get(date_str, {})
    today_total = day_data.get("总编辑数", 0)
    accumulative_total += today_total  # 更新总累计值

    col = 2  # 从第二列开始（命名空间列）
    for ns in sorted_ns:
        today_count = day_data.get(ns, 0)
        accumulative_counts[ns] += today_count  # 更新命名空间累计值

        # 计算并写入百分比
        if accumulative_total > 0:
            percentage = accumulative_counts[ns] / accumulative_total
        else:
            percentage = 0

        cell = ws.cell(row=row, column=col, value=percentage)
        cell.style = "percent_style"  # 应用百分比样式
        col += 1

    # 保留总编辑数（绝对值）
    ws.cell(row=row, column=col, value=accumulative_total)

    row += 1
    cur_date += one_day

# 如果表格数据过多，Excel可能不允许插入折线图，因此程序插入
chart = LineChart()
chart.title = "命名空间编辑占比趋势"
chart.style = 2
chart.y_axis.title = '占比'
chart.x_axis.title = '日期'
chart.height = 10  # 图表高度（英寸）
chart.width = 20   # 图表宽度（英寸）

# 数据范围：从第2行开始，第2列到倒数第2列（命名空间占比列）
data = Reference(ws, min_col=2, min_row=1, max_col=len(sorted_ns)+1, max_row=row-1)
cats = Reference(ws, min_col=1, min_row=2, max_row=row-1)

chart.add_data(data, titles_from_data=True)
chart.set_categories(cats)

# 将图表插入到表格右侧空白区域（例如从第2列 + len(sorted_ns) + 3 开始）
chart_anchor_col = len(sorted_ns) + 3  # 留出2列空白
chart_anchor_cell = f"{openpyxl.utils.get_column_letter(chart_anchor_col)}1"

ws.add_chart(chart, chart_anchor_cell)

current_time = datafile[-14:]
base.output(f"{username}-editintegration-{current_time}.xlsx", wb, "xlsx")

print(f"结果已保存至{username}-editintegration-{current_time}.xlsx")
