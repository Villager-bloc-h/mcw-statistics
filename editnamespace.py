import json
import sys
import openpyxl

import base

datafile = base.editperiod_get_config()

try:
    with open(f"{datafile}.json", "r", encoding="utf-8") as contribs_file:
        contribs_data = json.load(contribs_file)
except FileNotFoundError:
    print("指定的文件不存在！")
    input("按任意键退出")
    sys.exit(1)

username = contribs_data[0]["user"]

wb = openpyxl.Workbook()
ws = wb.active
ws['A1'] = "namespace"
ws['B1'] = "命名空间"
ws['C1'] = "编辑数"
ws['D1'] = "百分比"
ws.column_dimensions['B'].width = 20.00

edit_count = {}
total_edit = 0

# 统计命名空间名称和编辑数
for rec in contribs_data:
    ns = rec['ns']
    ns_name = rec['title'].split(':', 1)[0] if ':' in rec['title'] else '（主）'
    key = (ns, ns_name if ns_name != 'Minecraft' else '（主）')
    edit_count[key] = edit_count.get(key, 0) + 1

# 将命名空间名称和编辑数写入表格
row = 2
for (ns, ns_name), count in sorted(edit_count.items(), key=lambda kv: kv[0][0]):
    ws[f'A{row}'] = ns
    ws[f'B{row}'] = ns_name
    ws[f'C{row}'] = count
    total_edit += count
    row += 1

# 将百分比写入表格
row = 2
for (ns, ns_name), count in sorted(edit_count.items(), key=lambda kv: kv[0][0]):
    percentage = (count / total_edit) * 100
    ws[f'D{row}'] = f"{percentage:.2f}%"
    row += 1

ws.cell(row=row, column=2).value = "编辑总数"
ws.cell(row=row, column=3).value = total_edit
ws.cell(row=row, column=4).value = "100.00%"

current_time = datafile[-14:]
wb.save(f"{username}-editnamespace-{current_time}.xlsx")

print(f"结果已保存至{username}-editnamespace-{current_time}.xlsx")
input("按任意键退出")
