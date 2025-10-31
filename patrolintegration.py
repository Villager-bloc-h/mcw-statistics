import time
import sys
from datetime import datetime, timedelta

import openpyxl
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import NamedStyle

import base

leend, lestart = base.patrolintegration_get_config()

params = {
    "action": "query",
    "format": "json",
    "list": "logevents",
    "formatversion": 2,
    "leprop": "timestamp|user",
    "leaction": "patrol/patrol",
    "lelimit": "max",
}

if leend != "":
    params.update({"leend": leend})

if lestart != "":
    params.update({"lestart": lestart})

patrol_result = []
last_lecontinue = ""
loop_count = 0

current_time = datetime.now().strftime("%Y%m%d%H%M%S")

base.login()

print("启动成功", end='\n\n')

while True:
    time.sleep(3)

    if last_lecontinue != "":
        last_params = params.copy()
        last_params.update({"lecontinue": last_lecontinue})
    else:
        last_params = params

    patrol_data = base.get_data(last_params)

    if 'error' in patrol_data:
        print("API返回错误信息，请检查时间戳格式是否正确。")
        input("按回车键退出")
        sys.exit(1)

    if not patrol_data['query']['logevents']:
        print("API未返回有效数据，请检查时间范围是否正确。")
        input("按回车键退出")
        sys.exit(1)

    patrol_result.extend(patrol_data['query']['logevents'])

    loop_count += 1
    print(f"成功获取第{loop_count}组数据")

    if 'continue' not in patrol_data:
        break

    last_lecontinue = patrol_data['continue']['lecontinue']

patrol_result = patrol_result[::-1]

wb = openpyxl.Workbook()
ws = wb.active
date_style = NamedStyle(name="date_style", number_format="YYYY-MM-DD")
wb.add_named_style(date_style)
ws.column_dimensions['A'].width = 10.89
ws['A1'] = "日期"

start_timestamp = base.extract_from_timestamp(patrol_result[0]["timestamp"])
end_timestamp = base.extract_from_timestamp(patrol_result[-1]["timestamp"])

patrol_count = {}
users = set()

for item in patrol_result:
    user = item["user"]
    users.add(user)
    date_str = base.extract_from_timestamp(item["timestamp"]).strftime("%Y-%m-%d")
    if date_str not in patrol_count:
        patrol_count[date_str] = {}
    patrol_count[date_str][user] = patrol_count[date_str].get(user, 0) + 1

sorted_users = sorted(list(users))

for i, user in enumerate(sorted_users):
    col = openpyxl.utils.get_column_letter(i + 2)
    ws.column_dimensions[col].width = 15
    ws.cell(row=1, column=i + 2, value=user)

row = 2
one_day = timedelta(days=1)
cur_timestamp = start_timestamp
while cur_timestamp <= end_timestamp:
    date_str = cur_timestamp.strftime("%Y-%m-%d")

    ws.cell(row=row, column=1, value=cur_timestamp).style = "date_style"

    for i, user in enumerate(sorted_users):
        user_daily_count = 0
        if date_str in patrol_count and user in patrol_count[date_str]:
            user_daily_count = patrol_count[date_str][user]

        if row == 2:
            user_total_value = user_daily_count
        else:
            user_total_value = ws.cell(row=row - 1, column=i + 2).value + user_daily_count
        ws.cell(row=row, column=i + 2, value=user_total_value)

    row += 1
    cur_timestamp += one_day

# 创建折线图
chart = LineChart()
chart.title = "巡查图"
chart.style = 2
chart.y_axis.title = '巡查数'
chart.x_axis.title = '日期'
chart.height = 10
chart.width = 20

# 数据范围
data = Reference(ws, min_col=2, min_row=1, max_col=len(sorted_users) + 1, max_row=row - 1)
cats = Reference(ws, min_col=1, min_row=2, max_row=row - 1)

chart.add_data(data, titles_from_data=True)
chart.set_categories(cats)

# 将图表插入到表格右侧
chart_anchor_col = len(sorted_users) + 3
chart_anchor_cell = f"{openpyxl.utils.get_column_letter(chart_anchor_col)}1"
ws.add_chart(chart, chart_anchor_cell)

wb.save(f"patrolintegration-{current_time}.xlsx")

print(f"结果已保存至patrolintegration-{current_time}.xlsx")
